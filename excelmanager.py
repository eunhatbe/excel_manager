
from datetime import datetime

# openpyxl 2.6.1
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image


class ExcelManager():
    def __init__(self):
        self.product_info_list = []

        self.base_url = None
        self.text_url = None
        self.save_url = None

        self.date = None

        self.product_name1 = None
        self.product_name2 = None
        self.product_price1 = None
        self.product_price2 = None

    def init_filepath(self, base_url, text_url, save_url):
        self.base_url = base_url
        self.text_url = text_url
        self.save_url = save_url

    def _check_path_exists(self):
        if not (self.base_url and self.text_url and self.save_url):
            return False
        return True

    def _read_datafile(self):

        if not self._check_path_exists():
            return

        self.image = Image(f'{self.base_url}/stamp.png')

        wb = openpyxl.load_workbook(f"{self.text_url}")
        ws = wb.active

        for i in range(2, ws.max_row + 1):
            product_info = {}
            product_info['name'] = ws[f'A{i}'].value
            product_info['quantity'] = ws[f'D{i}'].value
            product_info['delivery_count'] = int(ws[f'G{i}'].value / 5000)
            self.product_info_list.append(product_info)

    def init_date(self, date):
        # 만약 날짜가 지정되지 않았다면 오늘 날짜로 초기화
        if not date:
            self.date = datetime.today().strftime("%Y-%m-%d")
            return

        self.date = date

    def init_product(self,product_name1,product_price1,product_name2,product_price2):
        self.product_name1 = product_name1
        self.product_price1 = product_price1
        self.product_name2 = product_name2
        self.product_price2 = product_price2

    def run(self):

        self._read_datafile()

        # 워크북 생성
        wb = openpyxl.load_workbook(f"{self.base_url}/base.xlsx")

        for i in range(len(self.product_info_list)):
            ws = wb.active

            # 업체명
            ws['B5'] = self.product_info_list[i]['name']

            # 수량
            ws['G14'] = self.product_info_list[i]['quantity']
            ws['G15'] = self.product_info_list[i]['delivery_count']

            # 품명
            ws['C14'] = self.product_name1
            ws['C15'] = self.product_name2

            # 단가
            ws['H14'] = self.product_price1
            ws['H15'] = self.product_price2

            # 날짜
            ws['B3'] = self.date
            ws['B14'] = self.date[5:]
            ws['B15'] = self.date[5:]

            ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
            ws['G14'].alignment = Alignment(horizontal='center', vertical='center')
            ws['G15'].alignment = Alignment(horizontal='center', vertical='center')

            target_cell_j14 = ws['j14']  # 공급가액
            target_cell_j15 = ws['j15']  # 공급가액
            target_cell_i26 = ws['i26']  # 합계
            target_cell_b11 = ws['b11']  # 합계

            target_cell_j14.value = '=G14*H14'
            target_cell_j15.value = '=G15*H15'
            target_cell_i26.value = '=SUM(j14:j15)'
            target_cell_b11.value = '=SUM(j14:j15)'

            wb.save(f'{self.save_url}/{self.product_info_list[i]["name"]}{i}.xlsx')
