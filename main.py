
from datetime import datetime

# openpyxl 2.6.1
import openpyxl
from openpyxl.styles import Alignment

name_list = []
number_list1 = []
number_list2 = []

try:
    with open("text/name150.txt", 'r', encoding="utf-8") as f:
        file = f.read()
        name_list = file.split("\n")
except Exception as e:
    print(f"Error : {e}")

try:
    with open("text/number150.txt", 'r') as f:
        file = f.read()
        number_list1 = file.split("\n")
except Exception as e:
    print(f"Error : {e}")

try:
    with open("text/number2150.txt", 'r') as f:
        file = f.read()
        number_list2 = file.split("\n")
except Exception as e:
    print(f"Error : {e}")

# 워크북 생성
wb = openpyxl.load_workbook("base/base.xlsx")

# 필요 변수 생성
YMD = datetime.today().strftime("%Y-%m-%d")
MD = datetime.today().strftime("%m-%d")

for i in range(len(name_list)):

    ws = wb.active

    # 품명
    ws['B5'] = name_list[i]

    # 수량
    ws['G14'] = number_list1[i]
    ws['G15'] = number_list2[i]

    # 날짜
    ws['B3'] = YMD
    ws['B14'] = MD
    ws['B15'] = MD

    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G15'].alignment = Alignment(horizontal='center', vertical='center')

    target_cell_j14 = wb['in']['j14']   # 공급가액
    target_cell_j15 = wb['in']['j15']   # 공급가액
    target_cell_i26 = wb['in']['i26']   # 합계
    target_cell_b11 = wb['in']['b11']   # 합계

    target_cell_j14.value = '=G14*H14'
    target_cell_j15.value = '=G15*H15'
    target_cell_i26.value = '=SUM(j14:j15)'
    target_cell_b11.value = '=SUM(j14:j15)'

    wb.save(f'excel/{name_list[i]}.xlsx')


